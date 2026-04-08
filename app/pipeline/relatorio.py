"""
Gerador de relatórios Excel e CSV de saída.
"""
from __future__ import annotations

import csv
import io
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Cores de classificação
_COR_ALTA = "FFCCCC"    # vermelho claro
_COR_MEDIA = "FFF2CC"   # amarelo claro
_COR_BAIXA = "CCFFCC"   # verde claro

_HEADERS_ALERTAS = [
    "ID",
    "Tipo Ação",
    "Classificação",
    "Score Final",
    "Marca Base",
    "NCL Base",
    "Especificação Base",
    "Marca RPI",
    "NCL RPI",
    "Especificação RPI",
    "Processo RPI",
    "Despacho",
    "Titular RPI",
    "Camada",
    "Score Nome",
    "Score Fonético",
    "Score Spec",
    "Score Núcleo",
    "Score IA",
    "Justificativa IA",
    "Núcleo Base",
    "Núcleo RPI",
    "Classes Colidem",
    "Sigla?",
    "Desgastado?",
]


def _cor_classificacao(classificacao: str | None) -> PatternFill | None:
    mapa = {
        "ALTA": PatternFill(start_color=_COR_ALTA, end_color=_COR_ALTA, fill_type="solid"),
        "MEDIA": PatternFill(start_color=_COR_MEDIA, end_color=_COR_MEDIA, fill_type="solid"),
        "BAIXA": PatternFill(start_color=_COR_BAIXA, end_color=_COR_BAIXA, fill_type="solid"),
    }
    return mapa.get(classificacao or "")


def _despacho_label(codigo: str, nome: str) -> str:
    if codigo and nome:
        return f"{codigo} — {nome}"
    return codigo or nome or ""


def gerar_xlsx(
    resultados: list[dict],
    stats: dict,
    output_path: str,
) -> str:
    """
    Gera o relatório Excel com duas abas: Alertas e Resumo.
    Retorna o caminho do arquivo gerado.
    """
    wb = openpyxl.Workbook()

    # -----------------------------------------------------------------------
    # Aba Alertas
    # -----------------------------------------------------------------------
    ws_alertas = wb.active
    ws_alertas.title = "Alertas"

    # Header
    ws_alertas.append(_HEADERS_ALERTAS)
    for cell in ws_alertas[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for i, r in enumerate(resultados, start=1):
        classificacao = r.get("classificacao", "")
        tipo_acao = r.get("tipo_acao", "")
        tipo_label = "OPOSIÇÃO" if tipo_acao == "OPOSICAO" else "PAN"

        row_data = [
            i,
            tipo_label,
            classificacao,
            r.get("score_final") or r.get("score_nome", 0),
            r.get("marca_base", ""),
            r.get("ncl_base", ""),
            (r.get("spec_base", "") or "")[:300],
            r.get("marca_rpi", ""),
            r.get("ncl_rpi", ""),
            (r.get("spec_rpi", "") or "")[:300],
            r.get("processo_rpi", ""),
            _despacho_label(r.get("despacho_codigo", ""), r.get("despacho_nome", "")),
            r.get("titular_rpi", ""),
            r.get("camada_deteccao", ""),
            r.get("score_nome", ""),
            r.get("score_fonetico", ""),
            r.get("score_spec", ""),
            r.get("score_nucleo", ""),
            r.get("score_ia", ""),
            r.get("justificativa_ia", ""),
            r.get("nucleo_base", ""),
            r.get("nucleo_rpi", ""),
            "Sim" if r.get("classes_colidem_flag") else "Não",
            "Sim" if r.get("is_sigla") else "Não",
            "Sim" if r.get("is_desgastado") else "Não",
        ]
        ws_alertas.append(row_data)

        # Colorir célula de classificação (coluna 3)
        fill = _cor_classificacao(classificacao)
        if fill:
            ws_alertas.cell(row=i + 1, column=3).fill = fill

    # Auto-filtro e freeze
    ws_alertas.auto_filter.ref = ws_alertas.dimensions
    ws_alertas.freeze_panes = "A2"

    # Larguras das colunas
    larguras = [5, 12, 14, 10, 40, 8, 50, 40, 8, 50, 15, 40, 40, 8,
                10, 10, 10, 10, 10, 60, 30, 30, 12, 8, 10]
    for i, w in enumerate(larguras, start=1):
        ws_alertas.column_dimensions[get_column_letter(i)].width = w

    # -----------------------------------------------------------------------
    # Aba Resumo
    # -----------------------------------------------------------------------
    ws_resumo = wb.create_sheet("Resumo")

    resumo_data = [
        ["Relatório de Colidência de Marcas"],
        [],
        ["Data de execução", datetime.now().strftime("%d/%m/%Y %H:%M")],
        ["Número da RPI", stats.get("rpi_numero", "")],
        ["Data da RPI", stats.get("rpi_data", "")],
        [],
        ["VOLUMES"],
        ["Total carteira de clientes", stats.get("total_carteira", 0)],
        ["Total RPI analisada", stats.get("total_rpi", 0)],
        ["  → Marcas para OPOSIÇÃO", stats.get("total_rpi_oposicao", 0)],
        ["  → Marcas para PAN", stats.get("total_rpi_pan", 0)],
        [],
        ["ALERTAS POR TIPO DE AÇÃO"],
        ["OPOSIÇÃO (prazo 60 dias) — total", stats.get("alertas_oposicao", 0)],
        ["PAN (prazo 180 dias) — total", stats.get("alertas_pan", 0)],
        [],
        ["ALERTAS POR CLASSIFICAÇÃO"],
        ["ALTA", stats.get("alertas_alta", 0)],
        ["MÉDIA", stats.get("alertas_media", 0)],
        ["BAIXA", stats.get("alertas_baixa", 0)],
        ["TOTAL", stats.get("alertas_total", 0)],
        [],
        ["PIPELINE — VOLUME POR CAMADA"],
        ["Camada 1 (Nome idêntico)", stats.get("camada1_count", 0)],
        ["Camada 2 (Fonético)", stats.get("camada2_count", 0)],
        ["Camada 3 (Especificação)", stats.get("camada3_count", 0)],
        ["Camada 4 (Scoring)", stats.get("camada4_count", 0)],
    ]

    for row in resumo_data:
        ws_resumo.append(row)

    ws_resumo["A1"].font = Font(bold=True, size=14)
    for row in ws_resumo.iter_rows():
        for cell in row:
            if cell.row in (3, 7, 13, 17, 23) and cell.column == 1:
                cell.font = Font(bold=True)

    ws_resumo.column_dimensions["A"].width = 45
    ws_resumo.column_dimensions["B"].width = 20

    wb.save(output_path)
    return output_path


def gerar_csv_bytes(resultados: list[dict]) -> bytes:
    """Gera o CSV como bytes para download."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_HEADERS_ALERTAS)

    for i, r in enumerate(resultados, start=1):
        tipo_acao = r.get("tipo_acao", "")
        tipo_label = "OPOSIÇÃO" if tipo_acao == "OPOSICAO" else "PAN"

        writer.writerow([
            i,
            tipo_label,
            r.get("classificacao", ""),
            r.get("score_final") or r.get("score_nome", 0),
            r.get("marca_base", ""),
            r.get("ncl_base", ""),
            (r.get("spec_base", "") or "")[:300],
            r.get("marca_rpi", ""),
            r.get("ncl_rpi", ""),
            (r.get("spec_rpi", "") or "")[:300],
            r.get("processo_rpi", ""),
            _despacho_label(r.get("despacho_codigo", ""), r.get("despacho_nome", "")),
            r.get("titular_rpi", ""),
            r.get("camada_deteccao", ""),
            r.get("score_nome", ""),
            r.get("score_fonetico", ""),
            r.get("score_spec", ""),
            r.get("score_nucleo", ""),
            r.get("score_ia", ""),
            r.get("justificativa_ia", ""),
            r.get("nucleo_base", ""),
            r.get("nucleo_rpi", ""),
            "Sim" if r.get("classes_colidem_flag") else "Não",
            "Sim" if r.get("is_sigla") else "Não",
            "Sim" if r.get("is_desgastado") else "Não",
        ])

    return output.getvalue().encode("utf-8-sig")  # BOM para Excel abrir corretamente
