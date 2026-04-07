"""
Gera o relatório de colidências em formato Excel (padrão A Província).
"""

from __future__ import annotations

import os

import openpyxl
from openpyxl.styles import Alignment, Font


_HEADERS = [
    'PROCESSO CLIENTE',
    'MARCA CLIENTE',
    'CLASSE CLIENTE',
    'TITULAR CLIENTE',
    'PROCESSO TERCEIRO',
    'MARCA TERCEIRO',
    'CLASSE TERCEIRO',
]

_COL_WIDTHS = [15, 40, 15, 45, 15, 40, 15]


def generate_report(
    results: list[dict],
    rpi_numero: str,
    rpi_data: str,
    total_rpi: int,
    total_clientes: int,
    output_dir: str = '.',
) -> str:
    """
    Gera o arquivo Excel de colidências no formato padrão A Província.

    Retorna o caminho do arquivo gerado.

    Estrutura do arquivo:
        Row 1 : Título
        Row 2 : (vazia)
        Row 3 : "A PROVINCIA MARCAS E PATENTES"
        Row 4 : "RPI {numero} de {data}"
        Row 5 : "Total de marcas verificadas: {total_rpi}"
        Row 6 : "Total de marcas controladas: {total_clientes}"
        Row 7 : "Foram selecionadas {n} colidências nesta RPI"
        Row 8 : Header das colunas de dados
        Row 9+: Dados
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Colidências'

    # --- Cabeçalho informativo ---
    ws.append(['Relatório de Colidência de Marca'])
    ws.append([])
    ws.append(['A PROVINCIA MARCAS E PATENTES'])
    ws.append([f'RPI {rpi_numero} de {rpi_data}'])
    ws.append([f'Total de marcas verificadas: {total_rpi}'])
    ws.append([f'Total de marcas controladas: {total_clientes}'])
    ws.append([f'Foram selecionadas {len(results)} colidências nesta RPI'])

    # Estilo do título (row 1)
    ws['A1'].font = Font(bold=True, size=14)

    # --- Header das colunas (row 8) ---
    ws.append(_HEADERS)
    for cell in ws[8]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # --- Dados ---
    for r in results:
        ws.append([
            r['PROCESSO CLIENTE'],
            r['MARCA CLIENTE'],
            r['CLASSE CLIENTE'],
            r['TITULAR CLIENTE'],
            r['PROCESSO TERCEIRO'],
            r['MARCA TERCEIRO'],
            r['CLASSE TERCEIRO'],
        ])

    # --- Ajuste de largura das colunas ---
    for i, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # --- Nome do arquivo ---
    # Normalizar data para uso em nome de arquivo: "31/03/2026" → "31-03-2026"
    safe_data = rpi_data.replace('/', '-')
    filename = f"Relatorio_Colidencia_RPI_{rpi_numero}_{safe_data}_A_PROVINCIA.xlsx"
    output_path = os.path.join(output_dir, filename)

    wb.save(output_path)
    return output_path
